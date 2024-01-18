from fileinput import filename
import tkinter as tk
from tkinter.filedialog import askopenfilename
from tkinter import *
from tkinter import ttk
from tkinter import scrolledtext
import os
import glob
import re
from turtle import seth
import pandas as pd
from datetime import datetime
import mechanize
import wget
from encodings import utf_8
from openpyxl import Workbook
from tkinter.filedialog import askopenfile 
import threading


def download():
    #get list of miRs inputted, mechanize browser, input one by one, collect designated Genes
    print("miRNA name: %s" % (text_area.get("1.0", "end")))
    dataentry=text_area.get("1.0", "end")
    print("dataentry:", dataentry)

    miR=[]
    i=0
    tempstr=""
    dataentry=re.sub(" ", "", dataentry)
    dataentry=re.sub("[*]", "", dataentry)
    dataentry=re.sub(",", "", dataentry)
    dataentry=re.sub("\n", "~", dataentry)
    dataentry=re.sub("\r", "~", dataentry)
    dataentry=re.sub("~~", "~", dataentry)
    while i<len(dataentry):
        if dataentry[i]=='~':
            miR.append(tempstr)
            tempstr=""
        else:
            tempstr=tempstr+dataentry[i]
        i+=1
    miR.append(tempstr)
    miR=list(filter(None, miR))
    print(miR)
    br = mechanize.Browser() # initial browser
    br.set_handle_robots(False) # ignore robots
    br.set_handle_refresh(False)    # can sometimes hang without this
    br.addheaders = [('User-agent', 'Firefox')] # [('User-agent', 'Firefox')]
    species = {} # species hash table which saves species and related website part
    species['Human']='vert_80'
    species['Mouse']='mmu_80'
    for m in miR:
        for s in species:
            ul = "https://www.targetscan.org/" + species[s] + "/" # first query page url
            print (ul)
            br.open(ul) # open first query page
            #for form in br.forms():
            #    print (form)
            br.form = list(br.forms())[0] # when form is unnamed, otherwise br.select_form("formname") after checking br.forms() -> form.name
            br.form['mirg'] = m # fill the query form item gid with gene name g
            r = str(br.submit().read().decode("utf-8")) # submit query form and get feedback r
            db = [] # data base list
            if "is not in our miRNA database." in r: # case 1: miR m database does not exist
                print("\nmiRNA {} is not in the database".format(m))
            else: # case 2: miR m database exists then extract database information for each g
                db = re.findall(r"(target=new>[a-zA-Z0-9\-\.]*</a></td><td><a)", r)
            countd=0
            newdb=[]
            for d in db:
                d=re.sub("target=new>", "", d)
                d=re.sub("</a></td><td><a", "", d)
                countd+=1
                newdb.append(d)
            
            print(countd)
            print(newdb)
            writetoexcel(m, s, newdb)

def writetoexcel(miRname, speciesname, database):
    wb=Workbook()
    sheet2=wb.active
    #sheet2=wb.create_sheet()
    sheet2['A1']="miRNA"
    sheet2['B1']="Gene"
    for pos, val in enumerate(database):
        sheet2.cell(row=pos+2, column=2).value = val # 6 is the position col F
        sheet2.cell(row=pos+2, column=1).value=miRname
    #for  i in database:
    #    sheet2.append(i)
    date_time = datetime.now().strftime("%Y%m%d_%H%M%S") # current date and time
    savename=date_time+"_"+miRname+"_"+speciesname+"_FinalResults.xlsx"
    wb.save(savename)

def returnall():
    #compile all collected genes and create master excel
    print("in return")

    


def functionnest1():
    download()
    returnall()



root = tk.Tk()
root.geometry('1800x1200')
root.title("M2 Binding Analysis Gene Search")
ttk.Label(root, text="Enter miRs:",
          font=("Cambria", 15)).grid(column=0, row=1)

text_area = scrolledtext.ScrolledText(root, wrap=tk.WORD,
                                      width=35, height=10,
                                      font=("Cambria", 12))
  
text_area.grid(column=0, row=2, pady=10, padx=10)

# placing cursor in text area
text_area.focus()
label_18 =Label(root,text="Enter a single miRNA per row", width=40,font=("Cambria",10))
label_18.place(x=40,y=232)
btn = Button(root, text = 'Quit', command = root.destroy)
btn.place(x=100, y=280)
btn1 = Button(root, text = 'Start', command = threading.Thread(target=functionnest1).start)
btn1.place(x=200, y=280)

tk.mainloop()
